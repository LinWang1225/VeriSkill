import openpyxl
from collections import defaultdict
wb = openpyxl.load_workbook("/tmp/boe_mill.xlsx", read_only=True, data_only=True)
ws = None
for s in wb.sheetnames:
    if s.startswith("M15."):
        ws = wb[s]; break
d = defaultdict(list)
for row in ws.iter_rows(min_row=7, values_only=True):
    yr, mo, val = row[0], row[1], row[2]
    if yr is None: continue
    try: yi = int(yr)
    except: continue
    if yi in (1950,1951,1952) and val is not None:
        d[yi].append(float(val))
for yi in (1950,1951,1952):
    vals = d[yi]
    gbp_per_usd_monthly = [1.0/v for v in vals]
    mean_gbp_per_usd = sum(gbp_per_usd_monthly)/len(gbp_per_usd_monthly)
    print(yi, "n=", len(vals), "mean_gbp_per_usd=", mean_gbp_per_usd)
    print("   monthly_usd_per_gbp:", vals)
    print("   monthly_gbp_per_usd:", [round(v,8) for v in gbp_per_usd_monthly])
